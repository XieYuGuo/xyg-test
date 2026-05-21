#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 Excel 文件内容转换为 JSON 数组
表头作为 key，每行数据作为 value
输出文件名格式: excel2json_毫秒时间戳.json
"""

import json
import sys
import os
import time
import openpyxl


def excel2json(excel_path, sheet_name=None):
    """读取 Excel 文件并转为 JSON 数组"""
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"工作表 '{sheet_name}' 不存在，可用工作表: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print("Excel 数据不足（至少需要表头行和一行数据）")
        sys.exit(1)

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

    result = []
    for row in rows[1:]:
        # 跳过全空行
        if all(cell is None for cell in row):
            continue
        item = {}
        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None
            # 数值类型保持原样，其余转字符串
            if isinstance(val, (int, float)):
                item[header] = val
            elif val is None:
                item[header] = ""
            else:
                item[header] = str(val)
        result.append(item)

    return result


def main():
    if len(sys.argv) < 2:
        print("用法: python excel2json.py <excel文件路径> [工作表名称]")
        print("示例: python excel2json.py aaa.xlsx")
        print("      python excel2json.py aaa.xlsx Sheet1")
        sys.exit(1)

    excel_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    data = excel2json(excel_path, sheet_name)

    # 输出文件路径: 脚本所在目录 / excel2json_毫秒时间戳.json
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ts = int(time.time() * 1000)
    output_path = os.path.join(script_dir, f"excel2json_{ts}.json")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"转换完成，共 {len(data)} 条记录")
    print(f"输出文件: {output_path}")


if __name__ == "__main__":
    main()
